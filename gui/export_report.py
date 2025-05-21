import os
import sqlite3
import pandas as pd
from PyQt6.QtWidgets import QFileDialog, QMessageBox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys
def generate_report_to_excel():
    # Получить путь к базе данных
    if getattr(sys, 'frozen', False):
        base_path = os.path.dirname(sys.executable)
    else:
        # мы сейчас внутри gui/, поднимаемся на уровень выше
        base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    db_path =os.path.join(base_path, "computers.db")
    
    if not os.path.exists(db_path):
        QMessageBox.critical(None, "Ошибка", f"База данных не найдена по пути: {db_path}")
        return

    # Выбор файла для сохранения отчета
    file_dialog = QFileDialog()
    file_dialog.setAcceptMode(QFileDialog.AcceptMode.AcceptSave)
    file_dialog.setNameFilter("Excel Files (*.xlsx)")
    file_dialog.setDefaultSuffix("xlsx")
    if not file_dialog.exec():
        return

    filenames = file_dialog.selectedFiles()
    if not filenames:
        return
    filename = filenames[0]

    # Подключение к БД
    if getattr(sys, 'frozen', False):
        base_path = os.path.dirname(sys.executable)
    else:
        # мы сейчас внутри gui/, поднимаемся на уровень выше
        base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    db_path =os.path.join(base_path, "computers.db")
    print(db_path)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    try:
        cur.execute("SELECT id, arm_type, dst, device_type, serial_number, user, address, admin FROM computers")
    except sqlite3.OperationalError as e:
        QMessageBox.critical(None, "Ошибка базы данных", f"Ошибка при чтении таблицы: {e}")
        conn.close()
        return

    computers = cur.fetchall()
    rows = []

    for computer_id, arm_type, dst, device_type, serial, user, address, admin in computers:
        # Получить IP и MAC
        cur.execute("SELECT ip, mac, mask FROM network_interfaces WHERE computer_id = ?", (computer_id,))
        ip_data = cur.fetchall()
        mac = next((entry[1] for entry in ip_data if entry[1]), "")  # Правильный MAC
        mac_formatted = mac.upper()
        ip_10 = next((entry[0] for entry in ip_data if entry[0]), "")
        ip_11 = next((entry[0] for entry in ip_data if entry[0].startswith("11.")), "")

        row = [
            "",  # 1: пустой
            f"АРМ ИСОД 77#{mac_formatted}",  # 2
            arm_type or "",  # 3
            "",  # 4
            "",  # 5
            "ФКУ НПО \"СТиС\" МВД России",  # 6
            mac_formatted,  # 7
            ip_11,  # 8
            ip_10,  # 9
            "ИСОД МВД России",  # 10
            "Windows",  # 11
            dst or "",  # 12
            device_type or "",  # 13
            serial or "",  # 14
            user or "",  # 15
            "г. Москва",  # 16
            "Центральный аппарат МВД России",  # 17
            address or "",  # 18
            "ФКУ НПО \"СТиС\" МВД России",  # 19
            admin or ""  # 20
        ]
        rows.append(row)

    conn.close()

    df = pd.DataFrame(rows)
    try:
        df.to_excel(filename, index=False, header=False, startrow=0)

        # Автоширина столбцов
        wb = load_workbook(filename)
        ws = wb.active
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Номер колонки
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[get_column_letter(column)].width = max_length + 2
        wb.save(filename)

        QMessageBox.information(None, "Готово", f"Отчет успешно сохранен: {filename}")
    except Exception as e:
        QMessageBox.critical(None, "Ошибка сохранения", f"Не удалось сохранить файл: {e}")

if __name__ == "__main__":
    generate_report_to_excel()
